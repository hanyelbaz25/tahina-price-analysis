from __future__ import annotations

import io
import math
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import List, Tuple

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
DEFAULT_FILE = Path(__file__).with_name("مقارنة مشتريات جدة.xlsx")
APP_VERSION = "V4.4 Auto Months"

st.set_page_config(
    page_title=f"تحليل أسعار المشتريات {APP_VERSION}",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)
st.markdown(
    """
<style>
html,body,[class*=css]{direction:rtl;text-align:right}
.main .block-container{padding-top:.7rem;padding-left:1rem;padding-right:1rem;max-width:1600px}
.hero{background:linear-gradient(135deg,#17324d,#0e7490);color:#fff;border-radius:22px;padding:22px 28px;margin-bottom:14px;box-shadow:0 10px 28px rgba(14,116,144,.16)}
.hero h1{margin:0;font-size:31px;line-height:1.35}.hero p{margin:8px 0 0;opacity:.93}
div[data-testid=stMetric]{background:#fff;border:1px solid #e5e7eb;border-radius:16px;padding:14px;box-shadow:0 3px 14px rgba(0,0,0,.05);min-height:104px}
.note{background:#f8fafc;border-right:5px solid #0e7490;padding:12px 14px;border-radius:10px}
.decision{padding:14px;border-radius:12px;background:#f8fafc;border:1px solid #e2e8f0;margin:8px 0}
.badge{display:inline-block;padding:4px 10px;border-radius:999px;font-weight:700;font-size:13px}.red{background:#fee2e2;color:#991b1b}.green{background:#dcfce7;color:#166534}.yellow{background:#fef3c7;color:#92400e}.blue{background:#dbeafe;color:#1e40af}
[data-testid="stDataFrame"]{overflow-x:auto;border-radius:12px}
.mobile-toolbar{background:#ffffff;border:1px solid #e5e7eb;border-radius:16px;padding:12px 14px;margin-bottom:12px;box-shadow:0 3px 14px rgba(0,0,0,.04)}
.kpi-grid{display:grid;grid-template-columns:repeat(7,minmax(0,1fr));gap:10px;margin:8px 0 16px}
.kpi-card{background:#fff;border:1px solid #e5e7eb;border-radius:15px;padding:12px;text-align:center;box-shadow:0 3px 12px rgba(0,0,0,.045)}
.kpi-label{font-size:12px;color:#64748b;margin-bottom:6px}.kpi-value{font-size:21px;font-weight:800;color:#0f172a;overflow-wrap:anywhere}
[data-testid="stPlotlyChart"]{overflow:hidden}
button[kind="header"]{z-index:1000000}

/* تحسين عرض الهواتف */
@media (max-width: 768px){
  html,body{overflow-x:hidden!important}
  .main .block-container{padding:.55rem .55rem 1.2rem!important;max-width:100%!important}
  .hero{border-radius:14px;padding:15px 14px;margin-bottom:10px}
  .hero h1{font-size:21px!important;line-height:1.45}
  .hero p{font-size:13px;line-height:1.65}

  .kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr))!important;gap:8px!important}
  .kpi-card{padding:10px 8px!important;border-radius:12px!important}
  .kpi-label{font-size:11px!important}.kpi-value{font-size:18px!important}
  div[data-testid=stMetric]{min-height:auto!important;padding:10px 12px!important;border-radius:12px}
  div[data-testid=stMetric] label{font-size:12px!important}
  div[data-testid=stMetric] [data-testid="stMetricValue"]{font-size:22px!important}
  section[data-testid="stSidebar"]{display:none!important}

  /* منع شرائح الاختيار من تمديد الشاشة */
  [data-baseweb="select"]{max-width:100%!important}
  [data-baseweb="tag"]{max-width:94%!important;height:auto!important;white-space:normal!important}
  [data-baseweb="tag"] span{white-space:normal!important;overflow-wrap:anywhere!important}

  /* أزرار وقوائم بحجم لمس مناسب */
  .stButton button,.stDownloadButton button{width:100%!important;min-height:44px}
  div[role="radiogroup"] label{padding:.35rem 0!important}

  /* الرسوم والجداول */
  [data-testid="stPlotlyChart"]>div{width:100%!important}
  [data-testid="stDataFrame"]{max-width:100%!important;overflow-x:auto!important}
  [data-testid="stDataFrame"] iframe{min-width:620px!important}
  h1{font-size:24px!important} h2{font-size:20px!important} h3{font-size:17px!important}
  .note{font-size:12px;line-height:1.7;padding:10px}
}
</style>
""",
    unsafe_allow_html=True,
)


def col_index(ref: str) -> int:
    letters = re.match(r"[A-Z]+", ref).group(0)
    n = 0
    for ch in letters:
        n = n * 26 + ord(ch) - 64
    return n


def _cell_value(c, shared):
    typ = c.attrib.get("t")
    v = c.find("a:v", NS)
    if v is None or v.text is None:
        return None
    raw = v.text
    if typ == "s":
        return shared[int(raw)]
    if typ == "b":
        return raw == "1"
    try:
        return float(raw)
    except ValueError:
        return raw


def _month_key(label: str):
    text = str(label).strip().replace("شهر", "").strip()
    arabic = {
        "يناير": 1, "فبراير": 2, "مارس": 3, "أبريل": 4, "ابريل": 4,
        "مايو": 5, "يونيو": 6, "يوليو": 7, "أغسطس": 8, "اغسطس": 8,
        "سبتمبر": 9, "أكتوبر": 10, "اكتوبر": 10, "نوفمبر": 11, "ديسمبر": 12,
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }
    low = text.lower()
    for name, num in arabic.items():
        if name in low:
            y = re.search(r"20\d{2}", low)
            return (int(y.group()) if y else 0, num, low)
    nums = [int(x) for x in re.findall(r"\d+", low)]
    if len(nums) >= 2:
        year = next((x for x in nums if 2000 <= x <= 2100), 0)
        month = next((x for x in nums if 1 <= x <= 12 and x != year), 99)
        return (year, month, low)
    if nums and 1 <= nums[0] <= 12:
        return (0, nums[0], low)
    return (9999, 99, low)


def _infer_month_from_filename(filename: str) -> str | None:
    name = Path(filename).stem
    arabic_months = ["يناير", "فبراير", "مارس", "أبريل", "ابريل", "مايو", "يونيو", "يوليو", "أغسطس", "اغسطس", "سبتمبر", "أكتوبر", "اكتوبر", "نوفمبر", "ديسمبر"]
    for month in arabic_months:
        if month in name:
            year = re.search(r"20\d{2}", name)
            return f"{month} {year.group()}" if year else month
    ym = re.search(r"(20\d{2})[-_ ](0?[1-9]|1[0-2])", name)
    if ym:
        return f"{int(ym.group(2))}-{ym.group(1)}"
    my = re.search(r"(?:^|\D)(0?[1-9]|1[0-2])[-_ ](20\d{2})(?:\D|$)", name)
    if my:
        return f"{int(my.group(1))}-{my.group(2)}"
    return None


def read_xlsx_cached(source: bytes, filename: str = "ملف.xlsx") -> Tuple[pd.DataFrame, List[str]]:
    """Read all detected months from row 6. If no month columns exist, use column G and infer month from filename."""
    with zipfile.ZipFile(io.BytesIO(source)) as z:
        shared = []
        if "xl/sharedStrings.xml" in z.namelist():
            root = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for si in root.findall("a:si", NS):
                shared.append("".join((t.text or "") for t in si.findall(".//a:t", NS)))
        sheet_name = next((n for n in z.namelist() if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")), None)
        if not sheet_name:
            raise ValueError("لا توجد ورقة عمل صالحة داخل الملف.")
        sheet = ET.fromstring(z.read(sheet_name))
        rows = {}
        for row in sheet.findall(".//a:sheetData/a:row", NS):
            r = int(row.attrib["r"])
            rows[r] = {}
            for c in row.findall("a:c", NS):
                rows[r][col_index(c.attrib["r"])] = _cell_value(c, shared)

    header = rows.get(6, {})
    base = {2: "كود الصنف", 3: "اسم الصنف", 4: "التصنيف", 5: "الوحدة", 6: "المورد", 7: "السعر الأساسي"}
    month_cols = []
    for c in range(9, 250):
        h = header.get(c)
        if h is None or not str(h).strip():
            continue
        label = str(int(h)) if isinstance(h, float) and h.is_integer() else str(h).strip()
        # Accept numbers 1-12, Arabic/English month names, or labels containing a year/month.
        key = _month_key(label)
        if key[1] <= 12 or any(x in label.lower() for x in ["يناير","فبراير","مارس","أبريل","ابريل","مايو","يونيو","يوليو","أغسطس","اغسطس","سبتمبر","أكتوبر","اكتوبر","نوفمبر","ديسمبر","jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]):
            month_cols.append((c, label))

    inferred = None
    if not month_cols:
        inferred = _infer_month_from_filename(filename)
        if inferred:
            month_cols = [(7, inferred)]
    if not month_cols:
        raise ValueError("لم يتم العثور على شهور في الصف 6، ولم يمكن معرفة الشهر من اسم الملف.")

    recs = []
    for r in sorted(k for k in rows if k > 6):
        vals = rows[r]
        if vals.get(2) in (None, "") or vals.get(3) in (None, ""):
            continue
        rec = {label: vals.get(idx) for idx, label in base.items()}
        for idx, m in month_cols:
            try:
                x = float(vals.get(idx))
            except (TypeError, ValueError):
                x = math.nan
            rec[f"شهر {m}"] = x if pd.notna(x) and x > 0 else math.nan
        recs.append(rec)

    df = pd.DataFrame(recs)
    months = [f"شهر {m}" for _, m in month_cols]
    return df, months


def combine_workbooks(files: List[Tuple[str, bytes]]) -> Tuple[pd.DataFrame, List[str]]:
    """Combine one workbook with many months or several monthly workbooks. Later files override duplicate month values."""
    frames = []
    all_months = []
    for filename, content in files:
        frame, months = read_xlsx_cached(content, filename)
        frames.append((frame, months))
        all_months.extend(months)

    keys = ["كود الصنف", "اسم الصنف"]
    meta = ["التصنيف", "الوحدة", "المورد", "السعر الأساسي"]
    combined = None
    for frame, months in frames:
        frame = frame.copy()
        frame["كود الصنف"] = frame["كود الصنف"].astype(str).str.strip()
        frame["اسم الصنف"] = frame["اسم الصنف"].astype(str).str.strip()
        frame = frame.drop_duplicates(keys, keep="last").set_index(keys)
        if combined is None:
            combined = frame
            continue
        for col in meta + months:
            if col not in frame.columns:
                continue
            if col in combined.columns:
                combined[col] = frame[col].combine_first(combined[col])
            else:
                combined[col] = frame[col]
        for col in combined.columns:
            if col not in frame.columns:
                continue
        combined = combined.combine_first(frame)

    combined = combined.reset_index()
    month_unique = sorted(set(all_months), key=lambda x: _month_key(x))
    # Ensure all month columns exist after merging.
    month_unique = [m for m in month_unique if m in combined.columns]
    return combined, month_unique


def decision(row: pd.Series) -> str:
    change = row.get("التغير الكلي %", math.nan)
    risk = row.get("درجة المخاطر", 0)
    consecutive = row.get("أشهر ارتفاع متتالية", 0)
    if pd.isna(change):
        return "بيانات غير كافية"
    if change >= 15 or risk >= 70:
        return "بحث عن مورد بديل وتفاوض عاجل"
    if change >= 8 or consecutive >= 2:
        return "تفاوض لتثبيت السعر ومراجعة البدائل"
    if change >= 3:
        return "متابعة أسبوعية وطلب عرض سعر منافس"
    if change <= -7:
        return "فرصة شراء أو تثبيت السعر الحالي"
    return "استمرار الشراء مع المتابعة"


def enrich(df: pd.DataFrame, months: List[str]) -> pd.DataFrame:
    out = df.copy()
    for col in ["اسم الصنف", "التصنيف", "المورد", "الوحدة"]:
        out[col] = out[col].fillna("غير محدد").astype(str).str.strip()
    for m in months:
        out[m] = pd.to_numeric(out[m], errors="coerce")

    out["أول سعر"] = out[months].bfill(axis=1).iloc[:, 0]
    out["آخر سعر"] = out[months].ffill(axis=1).iloc[:, -1]
    out["التغير الكلي %"] = (out["آخر سعر"] / out["أول سعر"] - 1) * 100
    out.loc[out["أول سعر"].isna() | out["آخر سعر"].isna(), "التغير الكلي %"] = math.nan
    out["متوسط السعر"] = out[months].mean(axis=1)
    out["أعلى سعر"] = out[months].max(axis=1)
    out["أقل سعر"] = out[months].min(axis=1)
    out["التذبذب %"] = out[months].std(axis=1) / out["متوسط السعر"] * 100
    out["عدد الشهور المتاحة"] = out[months].count(axis=1)

    monthly_change_cols = []
    for i in range(1, len(months)):
        p, c = months[i - 1], months[i]
        col = f"تغير {p} ← {c} %"
        out[col] = (out[c] / out[p] - 1) * 100
        monthly_change_cols.append(col)

    def streak(r: pd.Series) -> int:
        vals = [r[c] for c in monthly_change_cols]
        count = 0
        for v in reversed(vals):
            if pd.notna(v) and v > 0.5:
                count += 1
            else:
                break
        return count

    out["أشهر ارتفاع متتالية"] = out.apply(streak, axis=1)
    out["الاتجاه"] = out["التغير الكلي %"].apply(
        lambda x: "مرتفع" if pd.notna(x) and x > 0.5 else ("منخفض" if pd.notna(x) and x < -0.5 else "مستقر")
    )
    out["درجة المخاطر"] = (
        out["التغير الكلي %"].abs().fillna(0) * 2.1
        + out["التذبذب %"].fillna(0) * 2.8
        + out["أشهر ارتفاع متتالية"] * 8
    ).clip(0, 100)
    out["مستوى المخاطر"] = out["درجة المخاطر"].apply(
        lambda x: "مرتفع" if x >= 60 else ("متوسط" if x >= 30 else "منخفض")
    )
    out["التوصية"] = out.apply(decision, axis=1)
    return out


def grouped(df: pd.DataFrame, key: str, months: List[str]) -> pd.DataFrame:
    rows = []
    for name, g in df.groupby(key, dropna=False):
        rows.append(
            {
                key: name if pd.notna(name) else "غير محدد",
                "عدد الأصناف": len(g),
                "متوسط التغير %": g["التغير الكلي %"].mean(),
                "متوسط التذبذب %": g["التذبذب %"].mean(),
                "مرتفعة": int((g["الاتجاه"] == "مرتفع").sum()),
                "منخفضة": int((g["الاتجاه"] == "منخفض").sum()),
                "عالية المخاطر": int((g["مستوى المخاطر"] == "مرتفع").sum()),
                "متوسط المخاطر": g["درجة المخاطر"].mean(),
                **{m: g[m].mean() for m in months},
            }
        )
    return pd.DataFrame(rows).sort_values("متوسط التغير %", ascending=False)


def pct(x) -> str:
    return "—" if pd.isna(x) else f"{x:,.2f}%"


def money(x) -> str:
    return "—" if pd.isna(x) else f"{x:,.2f} ر.س"


def direction_symbol(current, previous) -> str:
    if pd.isna(current) or pd.isna(previous):
        return "—"
    if current > previous + 1e-9:
        return "▲"
    if current < previous - 1e-9:
        return "▼"
    return "●"


def price_display(frame: pd.DataFrame, months: List[str]) -> pd.DataFrame:
    """Create a user-facing table with arrows beside monthly prices."""
    out = frame.copy()
    for i, m in enumerate(months):
        if i == 0:
            out[m] = out[m].apply(lambda v: "—" if pd.isna(v) else f"{v:,.2f}")
        else:
            prev = months[i - 1]
            out[m] = [
                "—" if pd.isna(cur) else f"{direction_symbol(cur, prv)} {cur:,.2f}"
                for cur, prv in zip(frame[m], frame[prev])
            ]
    return out


def style_price_table(display_df: pd.DataFrame, raw_df: pd.DataFrame, months: List[str]):
    def style_col(col):
        styles = []
        if col.name not in months or months.index(col.name) == 0:
            return ["" for _ in col]
        i = months.index(col.name)
        prev = months[i - 1]
        for cur, prv in zip(raw_df[col.name], raw_df[prev]):
            if pd.isna(cur) or pd.isna(prv):
                styles.append("")
            elif cur > prv + 1e-9:
                styles.append("background-color:#fee2e2;color:#b91c1c;font-weight:700")
            elif cur < prv - 1e-9:
                styles.append("background-color:#dcfce7;color:#166534;font-weight:700")
            else:
                styles.append("background-color:#fef3c7;color:#92400e;font-weight:700")
        return styles
    return display_df.style.apply(style_col, axis=0)


def make_excel(items: pd.DataFrame, cats: pd.DataFrame, sups: pd.DataFrame, months: List[str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.rightToLeft = True
    navy, teal, red, green, yellow, white = "17324D", "0E7490", "FECACA", "DCFCE7", "FEF3C7", "FFFFFF"
    thin = Side(style="thin", color="D1D5DB")

    ws.merge_cells("A1:H2")
    ws["A1"] = f"تقرير تحليل تغير أسعار المشتريات {APP_VERSION}"
    ws["A1"].font = Font(size=20, bold=True, color=white)
    ws["A1"].fill = PatternFill("solid", fgColor=teal)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    kpis = [
        ("إجمالي الأصناف", len(items)),
        ("التصنيفات", items["التصنيف"].nunique()),
        ("الموردون", items["المورد"].nunique()),
        ("متوسط التغير", items["التغير الكلي %"].mean() / 100),
        ("مرتفعة", int((items["الاتجاه"] == "مرتفع").sum())),
        ("عالية المخاطر", int((items["مستوى المخاطر"] == "مرتفع").sum())),
    ]
    for i, (label, val) in enumerate(kpis, 1):
        ws.cell(4, i, label)
        ws.cell(5, i, val)
        ws.cell(4, i).fill = PatternFill("solid", fgColor=navy)
        ws.cell(4, i).font = Font(color=white, bold=True)
        ws.cell(4, i).alignment = Alignment(horizontal="center")
        ws.cell(5, i).alignment = Alignment(horizontal="center")
        ws.cell(5, i).font = Font(size=15, bold=True)
    ws["D5"].number_format = "0.00%"

    top = items.nlargest(10, "التغير الكلي %")[["اسم الصنف", "التغير الكلي %"]]
    ws.append([])
    ws.append(["أعلى الأصناف ارتفاعًا", "نسبة التغير"])
    start = ws.max_row
    for row in top.itertuples(index=False):
        ws.append([row[0], row[1] / 100])
    for r in range(start + 1, ws.max_row + 1):
        ws.cell(r, 2).number_format = "0.00%"
    chart = BarChart()
    chart.title = "أعلى 10 أصناف ارتفاعًا"
    chart.add_data(Reference(ws, min_col=2, min_row=start, max_row=ws.max_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=ws.max_row))
    chart.height, chart.width = 8, 15
    ws.add_chart(chart, "D8")

    def add_sheet(name: str, frame: pd.DataFrame):
        sh = wb.create_sheet(name)
        sh.sheet_view.rightToLeft = True
        cols = list(frame.columns)
        sh.append(cols)
        for row in frame.itertuples(index=False, name=None):
            sh.append([None if pd.isna(v) else v for v in row])
        for cell in sh[1]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color=white, bold=True)
            cell.alignment = Alignment(horizontal="center")
        sh.freeze_panes = "A2"
        sh.auto_filter.ref = sh.dimensions
        for row in sh.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for i, col in enumerate(cols, 1):
            sh.column_dimensions[get_column_letter(i)].width = min(max(len(str(col)) + 2, 12), 32)
            if "%" in str(col):
                for r in range(2, sh.max_row + 1):
                    if isinstance(sh.cell(r, i).value, (int, float)):
                        sh.cell(r, i).value /= 100
                        sh.cell(r, i).number_format = "0.00%"
        for m in months:
            if m in cols:
                ci = cols.index(m) + 1
                if ci > 1:
                    prev = ci - 1
                    rng = f"{get_column_letter(ci)}2:{get_column_letter(ci)}{sh.max_row}"
                    sh.conditional_formatting.add(
                        rng,
                        CellIsRule(
                            operator="greaterThan",
                            formula=[f"{get_column_letter(prev)}2"],
                            fill=PatternFill("solid", fgColor=red),
                            font=Font(color="9B1C1C", bold=True),
                        ),
                    )
                    sh.conditional_formatting.add(
                        rng,
                        CellIsRule(
                            operator="lessThan",
                            formula=[f"{get_column_letter(prev)}2"],
                            fill=PatternFill("solid", fgColor=green),
                            font=Font(color="166534"),
                        ),
                    )
        if "درجة المخاطر" in cols:
            ci = cols.index("درجة المخاطر") + 1
            rng = f"{get_column_letter(ci)}2:{get_column_letter(ci)}{sh.max_row}"
            sh.conditional_formatting.add(rng, ColorScaleRule(start_type="num", start_value=0, start_color=green, mid_type="num", mid_value=50, mid_color=yellow, end_type="num", end_value=100, end_color=red))
        return sh

    base_cols = ["كود الصنف", "اسم الصنف", "التصنيف", "الوحدة", "المورد"]
    export_items = items[base_cols].copy()
    for i, m in enumerate(months):
        export_items[m] = items[m]
        if i > 0:
            prev = months[i - 1]
            export_items[f"حركة {m}"] = [direction_symbol(c, p) for c, p in zip(items[m], items[prev])]
    tail_cols = ["التغير الكلي %", "التذبذب %", "الاتجاه", "أشهر ارتفاع متتالية", "درجة المخاطر", "مستوى المخاطر", "التوصية"]
    for c in tail_cols:
        export_items[c] = items[c]
    data_sheet = add_sheet("البيانات", export_items)
    # Color arrows and monthly cells in the exported workbook.
    cols = list(export_items.columns)
    for i, m in enumerate(months):
        if i == 0:
            continue
        mci = cols.index(m) + 1
        aci = cols.index(f"حركة {m}") + 1
        prevci = cols.index(months[i - 1]) + 1
        for r in range(2, data_sheet.max_row + 1):
            cur = data_sheet.cell(r, mci).value
            prv = data_sheet.cell(r, prevci).value
            arrow = data_sheet.cell(r, aci)
            if isinstance(cur, (int, float)) and isinstance(prv, (int, float)):
                if cur > prv + 1e-9:
                    fill, color = PatternFill("solid", fgColor=red), "B91C1C"
                elif cur < prv - 1e-9:
                    fill, color = PatternFill("solid", fgColor=green), "166534"
                else:
                    fill, color = PatternFill("solid", fgColor=yellow), "92400E"
                data_sheet.cell(r, mci).fill = fill
                data_sheet.cell(r, mci).font = Font(color=color, bold=True)
                arrow.fill = fill
                arrow.font = Font(color=color, bold=True, size=14)
                arrow.alignment = Alignment(horizontal="center")
    add_sheet("تحليل التصنيفات", cats)
    add_sheet("تحليل الموردين", sups)
    rec = items[["اسم الصنف", "التصنيف", "المورد", "التغير الكلي %", "أشهر ارتفاع متتالية", "درجة المخاطر", "التوصية"]].sort_values("درجة المخاطر", ascending=False)
    add_sheet("مركز القرار", rec)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


st.markdown(
    f'<div class="hero"><h1>📊 برنامج تحليل تغير أسعار المشتريات {APP_VERSION}</h1><p>لوحة تنفيذية مع اكتشاف الشهور ودمج الملفات الشهرية تلقائيًا، وتحليل الصنف والتصنيف والمورد ومركز القرار.</p></div>',
    unsafe_allow_html=True,
)

st.markdown('<div class="mobile-toolbar"><b>📱 تنقل سريع وإعدادات مناسبة للهاتف</b></div>', unsafe_allow_html=True)

page = st.selectbox(
    "اختر الصفحة",
    ["🏠 لوحة المدير", "📦 تحليل الصنف", "🏷 تحليل التصنيف", "🚚 تحليل المورد", "📡 مراقبة الأسعار", "🎯 مركز القرار", "📑 البيانات والتصدير"],
    index=0,
)

with st.expander("⚙️ رفع الملفات والشهور التلقائية", expanded=False):
    uploaded_files = st.file_uploader(
        "اختر ملف Excel واحدًا أو عدة ملفات شهرية",
        type=["xlsx"],
        accept_multiple_files=True,
        help="يمكن رفع ملف واحد يحتوي كل الشهور، أو ملف مستقل لكل شهر. عند تكرار الشهر تؤخذ بيانات آخر ملف مرفوع.",
    )
    st.caption("يكتشف البرنامج جميع الشهور من الصف 6 تلقائيًا. وإذا كان الملف شهريًا بلا أعمدة شهور، اكتب الشهر في اسم الملف مثل: مشتريات_يونيو_2026.xlsx.")

try:
    file_inputs = [(f.name, f.getvalue()) for f in uploaded_files] if uploaded_files else [(DEFAULT_FILE.name, DEFAULT_FILE.read_bytes())]
    base, months = combine_workbooks(file_inputs)
    if len(months) < 2:
        raise ValueError("يلزم توفر شهرين على الأقل لإجراء المقارنة. ارفع ملف شهر إضافيًا.")
    data = enrich(base, months)
    st.success(f"تم اكتشاف {len(months)} شهرًا تلقائيًا: " + "، ".join(m.replace("شهر ", "") for m in months))
except Exception as exc:
    st.error(f"تعذر قراءة الملف: {exc}")
    st.stop()

cats_all = sorted(data["التصنيف"].unique(), key=str.casefold)
sups_all = sorted(data["المورد"].unique(), key=str.casefold)
with st.expander("🔎 الفلاتر والبحث", expanded=False):
    selected_cats = st.multiselect("التصنيفات", cats_all, default=cats_all)
    selected_sups = st.multiselect("الموردون", sups_all, default=sups_all)
    selected_risks = st.multiselect("مستوى المخاطر", ["مرتفع", "متوسط", "منخفض"], default=["مرتفع", "متوسط", "منخفض"])
    search = st.text_input("بحث باسم الصنف")

filtered = data[
    data["التصنيف"].isin(selected_cats)
    & data["المورد"].isin(selected_sups)
    & data["مستوى المخاطر"].isin(selected_risks)
].copy()
if search.strip():
    filtered = filtered[filtered["اسم الصنف"].str.contains(search.strip(), case=False, na=False)]
if filtered.empty:
    st.warning("لا توجد بيانات تطابق الفلاتر الحالية.")
    st.stop()

cat_sum = grouped(filtered, "التصنيف", months)
sup_sum = grouped(filtered, "المورد", months)

kpis = [
    ("الأصناف", len(filtered)),
    ("التصنيفات", filtered["التصنيف"].nunique()),
    ("الموردون", filtered["المورد"].nunique()),
    ("▲ مرتفعة", int((filtered["الاتجاه"] == "مرتفع").sum())),
    ("▼ منخفضة", int((filtered["الاتجاه"] == "منخفض").sum())),
    ("متوسط التغير", pct(filtered["التغير الكلي %"].mean())),
    ("عالية المخاطر", int((filtered["مستوى المخاطر"] == "مرتفع").sum())),
]
kpi_html = '<div class="kpi-grid">' + ''.join(
    f'<div class="kpi-card"><div class="kpi-label">{label}</div><div class="kpi-value">{value}</div></div>'
    for label, value in kpis
) + '</div>'
st.markdown(kpi_html, unsafe_allow_html=True)

if page == "🏠 لوحة المدير":
    a, b = st.columns(2)
    with a:
        top = filtered.nlargest(10, "التغير الكلي %").sort_values("التغير الكلي %")
        fig_top = px.bar(top, x="التغير الكلي %", y="اسم الصنف", orientation="h", title="أعلى 10 أصناف ارتفاعًا", text_auto=".2f")
        fig_top.update_traces(marker_color="#dc2626")
        st.plotly_chart(fig_top, use_container_width=True)
    with b:
        low = filtered.nsmallest(10, "التغير الكلي %").sort_values("التغير الكلي %", ascending=False)
        fig_low = px.bar(low, x="التغير الكلي %", y="اسم الصنف", orientation="h", title="أعلى 10 أصناف انخفاضًا", text_auto=".2f")
        fig_low.update_traces(marker_color="#16a34a")
        st.plotly_chart(fig_low, use_container_width=True)
    a, b = st.columns(2)
    with a:
        trend = filtered[months].mean().reset_index()
        trend.columns = ["الشهر", "متوسط السعر"]
        st.plotly_chart(px.line(trend, x="الشهر", y="متوسط السعر", markers=True, title="اتجاه متوسط الأسعار"), use_container_width=True)
    with b:
        dist = filtered["الاتجاه"].value_counts().rename_axis("الاتجاه").reset_index(name="العدد")
        st.plotly_chart(px.pie(dist, names="الاتجاه", values="العدد", hole=.5, title="توزيع الاتجاهات"), use_container_width=True)
    st.subheader("أكثر التصنيفات تأثرًا")
    st.dataframe(cat_sum.head(10), use_container_width=True, hide_index=True)

elif page == "📦 تحليل الصنف":
    names = sorted(filtered["اسم الصنف"].unique(), key=str.casefold)
    name = st.selectbox("اختر الصنف", names)
    row = filtered[filtered["اسم الصنف"] == name].iloc[0]
    cols = st.columns(7)
    values = [
        ("أول سعر", money(row["أول سعر"])), ("آخر سعر", money(row["آخر سعر"])), ("التغير", pct(row["التغير الكلي %"])),
        ("المتوسط", money(row["متوسط السعر"])), ("التذبذب", pct(row["التذبذب %"])),
        ("ارتفاع متتالٍ", int(row["أشهر ارتفاع متتالية"])), ("المخاطر", f"{row['درجة المخاطر']:.1f}/100"),
    ]
    for c, (l, v) in zip(cols, values):
        c.metric(l, v)
    p = pd.DataFrame({"الشهر": months, "السعر": [row[m] for m in months]})
    p["التغير الشهري %"] = p["السعر"].pct_change() * 100
    a, b = st.columns(2)
    with a:
        st.plotly_chart(px.line(p, x="الشهر", y="السعر", markers=True, title=f"حركة سعر {name}"), use_container_width=True)
    with b:
        fig_change = px.bar(p, x="الشهر", y="التغير الشهري %", title="التغير الشهري %", text_auto=".2f")
        fig_change.update_traces(marker_color=["#9ca3af" if pd.isna(v) else ("#dc2626" if v > 0 else ("#16a34a" if v < 0 else "#d97706")) for v in p["التغير الشهري %"]])
        st.plotly_chart(fig_change, use_container_width=True)
    st.info(f"التوصية: {row['التوصية']}")

elif page == "🏷 تحليل التصنيف":
    cat = st.selectbox("اختر التصنيف", sorted(filtered["التصنيف"].unique(), key=str.casefold))
    group = filtered[filtered["التصنيف"] == cat]
    a, b = st.columns(2)
    with a:
        trend = group[months].mean().reset_index(); trend.columns = ["الشهر", "متوسط السعر"]
        st.plotly_chart(px.line(trend, x="الشهر", y="متوسط السعر", markers=True, title=f"اتجاه تصنيف {cat}"), use_container_width=True)
    with b:
        st.plotly_chart(px.bar(group.sort_values("التغير الكلي %"), x="التغير الكلي %", y="اسم الصنف", orientation="h", title="تغير أصناف التصنيف", text_auto=".2f"), use_container_width=True)
    st.dataframe(group[["اسم الصنف", "المورد", "التغير الكلي %", "درجة المخاطر", "التوصية"]].sort_values("التغير الكلي %", ascending=False), use_container_width=True, hide_index=True)

elif page == "🚚 تحليل المورد":
    supplier = st.selectbox("اختر المورد", sorted(filtered["المورد"].unique(), key=str.casefold))
    group = filtered[filtered["المورد"] == supplier]
    cols = st.columns(6)
    values = [
        ("عدد الأصناف", len(group)), ("متوسط التغير", pct(group["التغير الكلي %"].mean())),
        ("مرتفعة", int((group["الاتجاه"] == "مرتفع").sum())), ("منخفضة", int((group["الاتجاه"] == "منخفض").sum())),
        ("عالية المخاطر", int((group["مستوى المخاطر"] == "مرتفع").sum())), ("متوسط المخاطر", f"{group['درجة المخاطر'].mean():.1f}"),
    ]
    for c, (l, v) in zip(cols, values):
        c.metric(l, v)
    st.plotly_chart(px.bar(group.sort_values("التغير الكلي %"), x="التغير الكلي %", y="اسم الصنف", orientation="h", title=f"أداء المورد: {supplier}", text_auto=".2f"), use_container_width=True)
    st.dataframe(group[["اسم الصنف", "التصنيف", "التغير الكلي %", "أشهر ارتفاع متتالية", "درجة المخاطر", "التوصية"]].sort_values("درجة المخاطر", ascending=False), use_container_width=True, hide_index=True)

elif page == "📡 مراقبة الأسعار":
    st.subheader("الأصناف التي ارتفعت في آخر شهر")
    last_change = f"تغير {months[-2]} ← {months[-1]} %"
    watch = filtered[pd.to_numeric(filtered[last_change], errors="coerce") > 0.5].copy().sort_values(last_change, ascending=False)
    c1, c2, c3 = st.columns(3)
    c1.metric("ارتفعت في آخر شهر", len(watch))
    c2.metric("أكثر ارتفاع شهري", pct(watch[last_change].max() if not watch.empty else math.nan))
    c3.metric("ارتفاع شهرين متتاليين", int((filtered["أشهر ارتفاع متتالية"] >= 2).sum()))
    watch_cols = ["اسم الصنف", "التصنيف", "المورد", months[-2], months[-1], last_change, "أشهر ارتفاع متتالية", "التوصية"]
    watch_raw = watch[watch_cols].reset_index(drop=True)
    watch_shown = price_display(watch_raw, months[-2:])
    st.dataframe(style_price_table(watch_shown, watch_raw, months[-2:]), use_container_width=True, hide_index=True)
    if not watch.empty:
        fig_watch = px.bar(watch.head(20).sort_values(last_change), x=last_change, y="اسم الصنف", orientation="h", title="أعلى الارتفاعات في آخر شهر", text_auto=".2f")
        fig_watch.update_traces(marker_color="#dc2626")
        st.plotly_chart(fig_watch, use_container_width=True)

elif page == "🎯 مركز القرار":
    urgent = filtered[filtered["التوصية"].str.contains("تفاوض|بديل", regex=True)].sort_values("درجة المخاطر", ascending=False)
    st.subheader("أصناف تحتاج تدخلًا")
    st.dataframe(urgent[["اسم الصنف", "التصنيف", "المورد", "التغير الكلي %", "أشهر ارتفاع متتالية", "درجة المخاطر", "التوصية"]], use_container_width=True, hide_index=True)
    st.subheader("محاكاة التفاوض")
    a, b, c = st.columns(3)
    discount = a.slider("نسبة الخصم المقترحة %", 0, 20, 5)
    qty = b.number_input("الكمية الافتراضية لكل صنف", min_value=1, value=100)
    selected_supplier = c.selectbox("المورد المستهدف", ["كل الموردين"] + sorted(urgent["المورد"].unique(), key=str.casefold))
    scenario = urgent if selected_supplier == "كل الموردين" else urgent[urgent["المورد"] == selected_supplier]
    saving = ((scenario["آخر سعر"] * discount / 100) * qty).sum()
    st.metric("الوفر التقديري", money(saving))
    st.caption("الوفر تقديري لعدم توفر كميات الشراء الفعلية في الملف؛ يُحسب على الكمية الافتراضية المدخلة.")

elif page == "📑 البيانات والتصدير":
    display = ["كود الصنف", "اسم الصنف", "التصنيف", "الوحدة", "المورد"] + months + [
        "التغير الكلي %", "التذبذب %", "الاتجاه", "أشهر ارتفاع متتالية", "درجة المخاطر", "مستوى المخاطر", "التوصية"
    ]
    raw_view = filtered[display].reset_index(drop=True)
    shown = price_display(raw_view, months)
    st.caption("▲ ارتفاع — أحمر | ▼ انخفاض — أخضر | ● ثبات — أصفر")
    st.dataframe(style_price_table(shown, raw_view, months), use_container_width=True, height=570, hide_index=True)
    xlsx = make_excel(filtered, cat_sum, sup_sum, months)
    st.download_button("📥 تحميل تقرير Excel الاحترافي", xlsx, f"تقرير_تحليل_الأسعار_{APP_VERSION}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

st.markdown('<div class="note"><b>إضافة شهر جديد:</b> ارفع ملفًا جديدًا من قسم «رفع الملفات والشهور التلقائية». يمكن رفع عدة ملفات معًا، ويُرتب البرنامج الشهور تلقائيًا.<br><b>منهجية الحساب:</b> نسبة التغير = (آخر سعر متاح ÷ أول سعر متاح − 1) × 100، مع استبعاد الأسعار الصفرية. مؤشر المخاطر يجمع التغير والتذبذب والارتفاعات المتتالية.</div>', unsafe_allow_html=True)
